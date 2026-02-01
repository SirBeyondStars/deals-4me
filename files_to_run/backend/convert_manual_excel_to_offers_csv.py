# files_to_run/backend/convert_manual_excel_to_offers_csv.py
# Purpose:
#   Convert a "raw" manual Excel export into a clean offers CSV for Supabase merge/export.
#
# Supports:
#   - Wegmans-style raw export (single column, lots of lines like "Loyalty discount price is:$13.00/ea")
#   - Simple "raw list" exports (price line, item line, optional "Price with membership" line)
#
# Output columns (standard D4M):
#   item_name, store, region, week_code,
#   promo_start, promo_end,
#   percent_off_prime, percent_off_nonprime,
#   sale_price,
#   manual_review_reason, source_file
#
# Notes:
#   - This script does NOT need the sheet to be "pretty".
#   - It scans raw text and extracts offer rows.

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook


STANDARD_HEADERS = [
    "item_name",
    "store",
    "region",
    "week_code",
    "promo_start",
    "promo_end",
    "percent_off_prime",
    "percent_off_nonprime",
    "sale_price",
    "manual_review_reason",
    "source_file",
]


def _norm_text(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # normalize weird non-breaking spaces
    s = s.replace("\u00a0", " ")
    return s.strip()


def read_all_cell_texts_xlsx(xlsx_path: Path) -> List[str]:
    wb = load_workbook(xlsx_path, data_only=True)
    texts: List[str] = []
    for ws in wb.worksheets:
        # iterate through *all* cells because some exports end up in weird columns
        for row in ws.iter_rows(values_only=True):
            for v in row:
                t = _norm_text(v)
                if t:
                    texts.append(t)
    return texts


# ----------------------------
# Wegmans raw export parser
# ----------------------------

_STOP_EXACT = {
    "Wegmans",
    "NEW ITEM",
    "Top",
    "Shopping",
    "Check Store",
    "Add to Cart",
    "Browse In-Store",
    "Weekly Sales",
    "Catering",
    "Order Online",
    "Grocery Pickup & Delivery",
}
# common nav/footer junk (seen in other sheets too)
_STOP_CONTAINS = [
    "Opens in a new tab",
    "About Whole Foods Market",
    "Privacy Notice",
    "Conditions of Use",
    "Site Map",
    "Corporate Policies",
    "Connect With Us",
    "Visit customer care",
]

_SIZE_RE = re.compile(
    r"^\s*[\d\.]+\s*(?:oz|ounce|ounces|fl\.?\s*oz|lb|pound|pounds|ct|count|g|kg|ml|l|liter|litre|pk|pack)\b",
    re.IGNORECASE,
)


def _is_noise_line(s: str) -> bool:
    if s in _STOP_EXACT:
        return True
    if any(x in s for x in _STOP_CONTAINS):
        return True
    low = s.lower()
    if low.startswith("original price was:"):
        return True
    if low.startswith("unit price is:"):
        return True
    if low.startswith("save "):
        return True
    if low.startswith("quantity:"):
        return True
    if "price with membership" in low:
        return True
    if _SIZE_RE.match(s):
        return True
    # pure $price lines are not product names
    if re.fullmatch(r"\$[\d]+(?:\.\d+)?(?:/\w+)?", s):
        return True
    return False


def parse_wegmans_loyalty_lines(texts: List[str]) -> List[Tuple[str, float]]:
    """
    Finds lines like: "Loyalty discount price is:$13.00/ea"
    and uses the nearest previous non-noise line as item_name.
    """
    recent: List[str] = []
    items: List[Tuple[str, float]] = []

    for s in texts:
        if "Loyalty discount price is:" in s:
            m = re.search(r"\$([\d]+(?:\.\d+)?)", s)
            if not m:
                continue
            price = float(m.group(1))

            # pick the most recent candidate line as name
            name = None
            for cand in reversed(recent[-15:]):
                if not _is_noise_line(cand) and re.search(r"[A-Za-z]", cand):
                    name = cand
                    break

            if name:
                items.append((name, price))
            continue

        recent.append(s)

    # de-dupe exact duplicates
    out: List[Tuple[str, float]] = []
    seen = set()
    for name, price in items:
        key = (name, price)
        if key in seen:
            continue
        seen.add(key)
        out.append((name, price))
    return out


# ----------------------------
# Simple raw list parser
# ----------------------------

_PRICE_LINE_RE = re.compile(r"^\$?\d+(?:\.\d{2})?$")


def parse_simple_price_name_pairs(texts: List[str]) -> List[Tuple[str, float]]:
    """
    Pattern:
      $3.99
      Some Item Name
      (optional) Price with membership
    """
    out: List[Tuple[str, float]] = []
    i = 0
    while i < len(texts) - 1:
        s = texts[i].strip()
        if _PRICE_LINE_RE.match(s):
            price = float(s.replace("$", ""))
            name = texts[i + 1].strip()
            # skip if name looks like junk
            if name and not _is_noise_line(name):
                out.append((name, price))
            i += 2
            continue
        i += 1

    # de-dupe
    ded: List[Tuple[str, float]] = []
    seen = set()
    for name, price in out:
        key = (name, price)
        if key in seen:
            continue
        seen.add(key)
        ded.append(key)
    return ded


def detect_schema(texts: List[str]) -> str:
    """
    Decide which parser to use.
    """
    if any("Loyalty discount price is:" in t for t in texts):
        return "wegmans_loyalty"
    # fallback
    if any(_PRICE_LINE_RE.match(t.strip()) for t in texts):
        return "raw_list"
    return "unknown"


def write_offers_csv(
    out_csv: Path,
    rows: List[dict],
) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=STANDARD_HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path to .xlsx")
    ap.add_argument("--output", required=True, help="Path to output .csv")
    ap.add_argument("--store", required=True, help="store key, e.g. wegmans")
    ap.add_argument("--region", required=True, help="region, e.g. NE")
    ap.add_argument("--week", required=True, help="week_code, e.g. wk_20260111")
    args = ap.parse_args()

    xlsx_path = Path(args.input)
    out_csv = Path(args.output)

    texts = read_all_cell_texts_xlsx(xlsx_path)
    schema = detect_schema(texts)
    print(f"[convert] Detected schema: {schema} for {xlsx_path.name}")

    pairs: List[Tuple[str, float]] = []
    if schema == "wegmans_loyalty":
        pairs = parse_wegmans_loyalty_lines(texts)
    elif schema == "raw_list":
        pairs = parse_simple_price_name_pairs(texts)
    else:
        print("[convert] Could not detect a supported schema. Writing 0 rows.")
        pairs = []

    rows: List[dict] = []
    for name, price in pairs:
        rows.append(
            {
                "item_name": name,
                "store": args.store,
                "region": args.region,
                "week_code": args.week.strip(),
                "promo_start": "",
                "promo_end": "",
                "percent_off_prime": "",
                "percent_off_nonprime": "",
                "sale_price": price,
                "manual_review_reason": "",
                "source_file": xlsx_path.name,
            }
        )

    write_offers_csv(out_csv, rows)
    print(f"[convert] Wrote {len(rows)} row(s) -> {out_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
